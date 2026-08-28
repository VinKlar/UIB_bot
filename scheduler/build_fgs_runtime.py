import json
import re
from pathlib import Path
from openpyxl import load_workbook


def load_groups_from_excel(excel_path: str) -> list[str]:
    wb = load_workbook(excel_path, data_only=True)
    sheet_name = next(
        (name for name in wb.sheetnames if name.strip().startswith("МАХ")),
        None,
    )
    if not sheet_name:
        raise ValueError("В книге не найден лист, название которого начинается с 'МАХ'")
    ws = wb[sheet_name]

    groups = []

    # В файле группы начинаются с 4 строки
    for row in ws.iter_rows(min_row=4, values_only=True):
        group = row[0]

        if not group:
            continue

        group = str(group).strip()

        if group.endswith(".0"):
            group = group[:-2]

        groups.append(group)

    return groups


def build_fgs_from_jsons(
    excel_path: str,
    json_dir: str
) -> dict[str, str]:

    groups = load_groups_from_excel(excel_path)
    json_files = list(Path(json_dir).glob("*.json"))

    def normalize_identifier(value: str) -> str:
        return re.sub(r"[^0-9a-zа-яё]", "", str(value).casefold())

    schedule_index = {}
    for file_path in json_files:
        try:
            data = json.loads(file_path.read_text(encoding="utf-8"))
            schedule_group = data["meta"]["group"]
        except (OSError, KeyError, json.JSONDecodeError, TypeError):
            continue

        for filename_group in file_path.stem.split(","):
            schedule_index[normalize_identifier(filename_group)] = schedule_group

    fgs = {}

    for group in groups:
        schedule_group = schedule_index.get(normalize_identifier(group))
        if schedule_group:
            fgs[group] = schedule_group

    return fgs
