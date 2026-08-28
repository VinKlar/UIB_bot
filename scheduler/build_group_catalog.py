import argparse
import json
import re
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


CODE_FIXES = {
    ("23.04.04", "Эксплуатация транспортно-технологических машин и комплексов"): "23.04.03",
    ("44.03.05", "Профессиональное обучение (по отраслям)"): "44.03.04",
}


def clean_text(value) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def clean_group(value) -> str:
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return clean_text(value)


def clean_code(value, direction_name: str) -> str:
    if isinstance(value, (date, datetime)):
        value = value.strftime("%d.%m.%y")

    code = clean_text(value).replace(" ", "").rstrip(".")
    return CODE_FIXES.get((code, direction_name), code)


def build_catalog(excel_path: Path) -> dict:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    worksheet = workbook.active

    directions = OrderedDict()
    groups = OrderedDict()

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=4, values_only=True),
        start=4,
    ):
        group = clean_group(row[0])
        profile = clean_text(row[1])
        direction_name = clean_text(row[3])
        study_form = clean_text(row[4])

        if not group:
            continue

        code = clean_code(row[2], direction_name)
        if not all((profile, code, direction_name, study_form)):
            raise ValueError(f"Не заполнены обязательные поля в строке {row_number}")

        direction_key = (code, direction_name)
        direction = directions.setdefault(
            direction_key,
            {
                "id": f"direction_{len(directions) + 1}",
                "code": code,
                "name": direction_name,
                "profiles": OrderedDict(),
            },
        )

        profile_key = (profile, study_form)
        profile_item = direction["profiles"].setdefault(
            profile_key,
            {
                "id": f"profile_{len(direction['profiles']) + 1}",
                "name": profile,
                "study_form": study_form,
                "groups": [],
            },
        )
        profile_item["groups"].append(group)

        if group in groups:
            raise ValueError(f"Группа {group!r} встречается повторно (строка {row_number})")

        groups[group] = {
            "group": group,
            "direction_id": direction["id"],
            "direction_code": code,
            "direction_name": direction_name,
            "profile": profile,
            "study_form": study_form,
        }

    result_directions = []
    for direction in directions.values():
        direction["profiles"] = list(direction["profiles"].values())
        result_directions.append(direction)

    return {
        "source": excel_path.name,
        "directions": result_directions,
        "groups": groups,
    }


def main():
    parser = argparse.ArgumentParser(description="Собрать каталог групп из Excel")
    parser.add_argument("excel_path", type=Path)
    parser.add_argument("output_path", type=Path)
    args = parser.parse_args()

    catalog = build_catalog(args.excel_path)
    args.output_path.parent.mkdir(parents=True, exist_ok=True)
    args.output_path.write_text(
        json.dumps(catalog, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(
        f"Сохранено: {args.output_path} "
        f"({len(catalog['directions'])} направлений, {len(catalog['groups'])} групп)"
    )


if __name__ == "__main__":
    main()
